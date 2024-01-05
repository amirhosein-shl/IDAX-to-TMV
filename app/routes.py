import subprocess

# Install required packages
subprocess.call(['pip', 'install', '-r', 'requirements.txt'])

import pandas as pd
import numpy as np
import os
import itertools
pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.max_columns', 500)
import openpyxl
from openpyxl import load_workbook
import re
import PyPDF2
from datetime import datetime
from datetime import timedelta

def IDAX(count, TMV, time):
    print('XXXXXXX IDAX was called')
    # final_path = r'M:\22\1.22417.00 - Sequim Transportation Master Plan Update\Data and Information\Traffic Counts\Final Data\Final Data\TMCs'
    final_path = count
    filenames = os.listdir(final_path)
    filenames = [file for file in filenames if file.endswith(time + '.xlsx') or file.endswith(time + '.xls')]
    num = 0
    workbook = load_workbook(TMV)

    # load_workbook(r"C:\Users\amirs\Desktop\Template\TM Voimelumes - updated - Copy.xlsx")
    for file in filenames:
        globals()['df_' + file.split('.')[0]] = pd.read_excel(os.path.join(final_path, file), skiprows=24,skipfooter=45)
        globals()['df_' + file.split('.')[0]] = globals()['df_' + file.split('.')[0]].dropna(axis=1, how='all')
        globals()['df_' + file.split('.')[0]] = globals()['df_' + file.split('.')[0]].dropna(axis=0, how='all')
        
        Street_A = globals()['df_' + file.split('.')[0]].iloc[1,11]
        Street_B = globals()['df_' + file.split('.')[0]].iloc[1,3]
        if Street_A == 0:
            Street_A = globals()['df_' + file.split('.')[0]].iloc[1,16]
        if Street_B == 0:
            Street_B = globals()['df_' + file.split('.')[0]].iloc[1,7]
            
        if Street_A == 'Driveway':
            Street_A = globals()['df_' + file.split('.')[0]].iloc[1,16]
        if Street_B == 'Driveway':
            Street_B = globals()['df_' + file.split('.')[0]].iloc[1,7]
        
        if globals()['df_' + file.split('.')[0]].shape[1]>23:
            print(str(Street_A)+'/'+str(Street_B)+'is 5-leg intersection!')
            num+=11
            continue
            
        Street_EB = globals()['df_' + file.split('.')[0]].iloc[1,3]
        Street_WB = globals()['df_' + file.split('.')[0]].iloc[1,7]
        Street_NB = globals()['df_' + file.split('.')[0]].iloc[1,11]
        Street_SB = globals()['df_' + file.split('.')[0]].iloc[1,16]
                
        globals()['df1_' + file.split('.')[0]] = pd.read_excel(os.path.join(final_path, file), skiprows=43,skipfooter=30)
        globals()['df1_' + file.split('.')[0]] = globals()['df1_' + file.split('.')[0]].dropna(axis=1, how='all')
        globals()['df1_' + file.split('.')[0]] = globals()['df1_' + file.split('.')[0]].dropna(axis=0, how='all')
        
        
        pdf_file = open(os.path.join(final_path, file.split('.')[0]+'.pdf'), 'rb')
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        page = pdf_reader.pages[0]
        page_text = page.extract_text()

        def find_dates(text):
            pattern = r"\b\d{1,2}/\d{1,2}/\d{4}\b"
            dates = re.findall(pattern, text)
            return dates
        dates_found = find_dates(page_text)
        date = dates_found[0]
        
        
        numeric_values = globals()['df_' + file.split('.')[0]].iloc[:,-1].apply(pd.to_numeric, errors='coerce')
        max_value = numeric_values[~numeric_values.isna()].max()
        matching_rows = globals()['df_' + file.split('.')[0]][globals()['df_' + file.split('.')[0]].iloc[:,-1] == max_value]
        row_indices = matching_rows.index
        ph = globals()['df_' + file.split('.')[0]].iloc[row_indices[0]-4:row_indices[0],:]
        numeric_values = ph.iloc[:,-2].apply(pd.to_numeric, errors='coerce')
        max_value1 = numeric_values[~numeric_values.isna()].max()
        phf = np.round(max_value/(max_value1*4),2)
        
        # Create a time object
        t1 = matching_rows.iloc[0,1]

        # Convert time to datetime with a dummy date
        dummy_date = datetime(1900, 1, 1)
        dt = datetime.combine(dummy_date, t1)

        # Subtract 45 minutes
        updated_dt = dt - timedelta(minutes=45)

        # Extract the time from the updated datetime
        updated_t1 = updated_dt.time()

        formatted_time = updated_t1.strftime("%H:%M")

        time_obj = datetime.strptime(formatted_time, '%H:%M')
        time_converted = time_obj.strftime('%H%M')
        
        a = globals()['df_' + file.split('.')[0]].iloc[-3:-1,:]

        ped_WB = globals()['df1_' + file.split('.')[0]].iloc[-1,12]
        if ped_WB == 0:
            ped_WB = ""
        ped_EB = globals()['df1_' + file.split('.')[0]].iloc[-1,13]
        if ped_EB == 0:
            ped_EB = ""
        ped_SB = globals()['df1_' + file.split('.')[0]].iloc[-1,14]
        if ped_SB == 0:
            ped_SB = ""
        ped_NB = globals()['df1_' + file.split('.')[0]].iloc[-1,15]
        if ped_NB == 0:
            ped_NB = ""

        Bike_EB = globals()['df1_' + file.split('.')[0]].iloc[-1,7]
        if Bike_EB == 0:
            Bike_EB = ""
        Bike_WB = globals()['df1_' + file.split('.')[0]].iloc[-1,8]
        if Bike_WB == 0:
            Bike_WB = ""
        Bike_NB = globals()['df1_' + file.split('.')[0]].iloc[-1,9]
        if Bike_NB == 0:
            Bike_NB = ""
        Bike_SB = globals()['df1_' + file.split('.')[0]].iloc[-1,10]
        if Bike_SB == 0:
            Bike_SB = ""
        try:
            HVEB = a.iloc[1,3:7].sum()/a.iloc[0,3:7].sum()
        except:
            HVEB = ''
        try:
            HVWB = a.iloc[1,7:11].sum()/a.iloc[0,7:11].sum()
        except:
            HVWB = ''
        try:
            HVNB = a.iloc[1,12:16].sum()/a.iloc[0,12:16].sum()
        except:
            HVNB = ''
        try:
            HVSB = a.iloc[1,17:21].sum()/a.iloc[0,17:21].sum() 
        except:
            HVSB = ''

        ### EB
        if Street_EB == 0:
            EBU4 = ''
            EBL5 = ''
            EBT6 = ''
            EBR7 = ''
            NBL8 = ''
            SBR4 = ''
            WBT6 = ''
            
        else:
            if a.iloc[0,3] == 0:
                EBU4 = ''
            else:
                EBU4  = a.iloc[0,3]
            EBL5 = a.iloc[0,4]
            EBT6 = a.iloc[0,5]
            EBR7 = a.iloc[0,6]

        ### WB
        if Street_WB == 0:
            WBR5 = ''
            WBT6 = ''
            WBL7 = ''
            WBU8 = ''
            SBL4 = ''
            NBR8 = ''
            EBT6 = ''
            
        else:
            if a.iloc[0,7] == 0:
                WBU8 = ''
            else:
                WBU8 = a.iloc[0,7]
            WBR5 = a.iloc[0,10]
            if Street_EB  == 0:
                WBT6 = ''
            else:
                WBT6 = a.iloc[0,9]
            WBL7 = a.iloc[0,8]

        ### NB
        if Street_NB  == 0:
            NBU8 = ''
            NBL8 = ''
            NBT8 = ''
            NBR8 = ''
            SBT4 = ''
            WBL7 = ''
            EBR7 = ''
            
        else:
            if a.iloc[0,11] == 0:
                NBU8 = ''
            else:
                NBU8 = a.iloc[0,11]
            if Street_EB  == 0:
                NBL8 = ''
            else:
                NBL8 = a.iloc[0,13]
            NBT8 = a.iloc[0,14]
            if Street_WB  == 0:
                NBR8 = ''
            else:
                NBR8 = a.iloc[0,15]

        ### SB
        if Street_SB  == 0:
            SBU4 = ''
            SBL4 = ''
            SBT4 = ''
            SBR4 = ''
            NBT8 = ''
            WBR5 = ''
            EBL5 = ''
            
        else:
            if a.iloc[0,16] == 0:
                SBU4 = ''
            else:
                SBU4 = a.iloc[0,16]
            if Street_WB  == 0:
                SBL4 = ''
            else:
                SBL4 = a.iloc[0,18]
            if Street_NB  == 0:
                SBT4 = ''
            else:
                SBT4 = a.iloc[0,19]
            if Street_EB  == 0:
                SBR4 = ''
            else:
                SBR4 = a.iloc[0,20]
        try:    
            if a.iloc[0,4] + a.iloc[0,5] + a.iloc[0,6] == 0:
                EBU4 = ''
                EBL5 = ''
                EBT6 = ''
                EBR7 = ''
        except:
            continue
        
        try:
            if a.iloc[0,8] + a.iloc[0,9] + a.iloc[0,10] == 0:
                WBR5 = ''
                WBT6 = ''
                WBL7 = ''
                WBU8 = ''
        except:
            continue
        
        try:
            if a.iloc[0,13] + a.iloc[0,14] + a.iloc[0,15] == 0:
                NBU8 = ''
                NBL8 = ''
                NBT8 = ''
                NBR8 = ''
        except:
            continue
            
        try:
            if a.iloc[0,18] + a.iloc[0,19] + a.iloc[0,20] == 0:
                SBU4 = ''
                SBL4 = ''
                SBT4 = ''
                SBR4 = ''
        except:
            continue
        
        ls_keys = ['J23', 'J24', 'J27', 'J28', 'J29', 'N27', 'N28', 'N29', 'K30', 'L30', 'M30', 'M26', 'L26', 'K26', 
                'D31', 'G29', 'C28', 'D28', 'E28', 'F28', 'E25', 'Q27', 'R26', 'U27', 'R30', 'Q26', 'Q30', 'U26', 'U30',
                'J26', 'N30', 'J30', 'N26', 'T33']
        
        ls_keys = [f'{cell[:1]}{int(cell[1:]) + num}' for cell in ls_keys]
        num += 11
    #     print(ls_keys)
        ls_values = [Street_A, Street_B, EBL5, EBT6, EBR7, WBR5, WBT6, WBL7, NBL8, NBT8, NBR8, SBL4, SBT4, SBR4, 
                    time_converted, phf, HVEB, HVWB, HVNB, HVSB, date, ped_EB, ped_SB, ped_WB, ped_NB, Bike_SB, 
                    Bike_EB, Bike_WB, Bike_NB, EBU4, WBU8, NBU8, SBU4, 'Signal']
        
        cell_updates = {}
        
        i=0
        for j in ls_keys:
            cell_updates[j] = ls_values[i]
            i+=1
        
    #     print(cell_updates)
        
        sheet = workbook['PM Traffic Volume Summary']
        worksheet = workbook.active

        for cell, value in cell_updates.items():
            try:
                sheet[cell] = value
            except Exception as e:
                print(f"Error in cell: {cell} - {str(e)}")
        
    print('XXXXXXX before save')
    workbook.save(TMV)
    print('XXXXXXX after save')





#### Start of the routes.py
from app import app
from flask import render_template, request, redirect, url_for
# from IDAX_to_TMV_py import IDAX



@app.route('/')
def index():
    return render_template('IDAX_to_TMV.html')


@app.route('/submit', methods=['POST'])
def process():
    input1 = request.form['input1']
    input2 = request.form['input2']
    input3 = request.form['input3']



    # Run your Python script using the data
    IDAX(count=input1, TMV=input2, time=input3)

    print('got end of the submit file')

    # return render_template('result.html', result=result)

